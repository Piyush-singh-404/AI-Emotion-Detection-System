import cv2
import time
from datetime import datetime
from deepface import DeepFace
from openpyxl import Workbook, load_workbook
import os

# Excel setup
file_name = "emotion_log.xlsx"

if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Emotion", "Percentage"])
    wb.save(file_name)

cap = cv2.VideoCapture(0)

prev_time = 0

while True:
    ret, frame = cap.read()

    results = DeepFace.analyze(frame, actions=['emotion'], enforce_detection=False)

    face_count = len(results)

    for result in results:
        emotion = result['dominant_emotion']
        emotions = result['emotion']
        percentage = "{:.2f}".format(emotions[emotion])

        face = result['region']

        x = face['x']
        y = face['y']
        w = face['w']
        h = face['h']

        # Emotion colors + recommendation
        if emotion == "happy":
            color = (0, 255, 0)
            suggestion = "Keep smiling!"
        elif emotion == "sad":
            color = (255, 0, 0)
            suggestion = "Take rest"
        elif emotion == "angry":
            color = (0, 0, 255)
            suggestion = "Calm down"
        elif emotion == "neutral":
            color = (0, 255, 255)
            suggestion = "Stay focused"
        else:
            color = (255, 255, 255)
            suggestion = ""

        # Face rectangle
        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)

        # Label background
        cv2.rectangle(frame, (x, y-35), (x+w, y), color, -1)

        # Emotion text
        cv2.putText(frame, f"{emotion} {percentage}%", (x+5, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                    (0,0,0), 2)

        # Recommendation text
        cv2.putText(frame, suggestion, (x, y+h+25),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                    color, 2)

    # FPS
    current_time = time.time()
    fps = 1 / (current_time - prev_time) if current_time != prev_time else 0
    prev_time = current_time

    cv2.putText(frame, f"FPS: {int(fps)}", (20, 80),
                cv2.FONT_HERSHEY_SIMPLEX, 1,
                (255,255,255), 2)

    # Date & Time
    now = datetime.now()
    current_date = now.strftime("%d-%m-%Y")
    current_time_text = now.strftime("%H:%M:%S")

    cv2.putText(frame, f"{current_date} {current_time_text}", (20, 120),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                (255,255,255), 2)

    # Face count
    cv2.putText(frame, f"Faces: {face_count}", (20, 160),
                cv2.FONT_HERSHEY_SIMPLEX, 1,
                (255,255,255), 2)

    # Title
    cv2.putText(frame, "AI Emotion Detection System", (20,40),
                cv2.FONT_HERSHEY_SIMPLEX, 1,
                (255,255,255), 2)

    cv2.imshow("Emotion Detection", frame)

    key = cv2.waitKey(1) & 0xFF

    # Screenshot save
    if key == ord('s'):
        filename = now.strftime("screenshot_%H%M%S.png")
        cv2.imwrite(filename, frame)

    # Excel save
    if key == ord('e'):
        wb = load_workbook(file_name)
        ws = wb.active
        ws.append([current_date, current_time_text, emotion, percentage])
        wb.save(file_name)

    # Quit
    if key == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()